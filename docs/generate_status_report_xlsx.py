from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

header_fill = PatternFill("solid", fgColor="111118")
header_font = Font(bold=True, color="D4AF37", size=11)
title_font = Font(bold=True, size=14, color="111118")
thin = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

fills = {
    "COMPLETE": PatternFill("solid", fgColor="DCFCE7"),
    "NEAR COMPLETE": PatternFill("solid", fgColor="DBEAFE"),
    "IN PROGRESS": PatternFill("solid", fgColor="FEF3C7"),
    "WAITING": PatternFill("solid", fgColor="F3F4F6"),
}


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# Exact rows from BillVyapp-Project-Status-Report-2026-08-07.md §3 (lines 61–86)
phases = [
    (1, "UI/UX Design & Frontend Development", "COMPLETE", 100, "All V1 screens, workflows, and responsive UI completed."),
    (2, "Backend API Development", "COMPLETE", 100, "Core APIs for appointments, billing, customers, inventory, memberships, messaging, and notifications completed."),
    (3, "Database Design & Implementation", "COMPLETE", 100, "Schema finalized with Prisma & MySQL (including expense delete-approval migrations)."),
    (4, "Authentication & User Management", "NEAR COMPLETE", 95, "Login, RBAC, and WhatsApp OTP completed. SMTP pending for email verification and password reset."),
    (5, "Customer Management Module", "COMPLETE", 100, "Customer CRUD, filters, history, visits, memberships, wallet, and coupon WhatsApp send completed."),
    (6, "Appointment Management", "COMPLETE", 100, "Walk-in, booking, calendar, appointment lifecycle, and scheduling completed."),
    (7, "Billing & POS Module", "NEAR COMPLETE", 95, "Billing, GST, discounts, split payments, invoices, memberships, and wallets completed. Razorpay pending; UPI QR + manual confirmation in use."),
    (8, "Inventory Management", "COMPLETE", 100, "Categories, vendors, products, purchase orders, stock adjustments, and usage logs completed."),
    (9, "Membership & Packages", "COMPLETE", 100, "Membership plans, packages, wallet, and service benefits implemented."),
    (10, "Notifications Module", "NEAR COMPLETE", 90, "In-app notifications + WhatsApp (Sparklebot via BullMQ/Memurai) completed. SMS and Email pending provider credentials."),
    (11, "Dashboard & Analytics", "COMPLETE", 100, "Revenue, appointments, customers, inventory, and business KPIs implemented (rbac + admin-dashboard)."),
    (12, "Playwright Automation Testing", "IN PROGRESS", 97, "Major business workflows automated and validated. Browser compatibility validation remaining."),
    (13, "Production Deployment (Cloudflare)", "COMPLETE", 100, "Deployed and live at https://billvyapp.com (Windows EC2 + Cloudflare)."),
    (14, "Production Verification", "IN PROGRESS", 92, "Core production verification in progress. WhatsApp path validating; final sign-off after SMTP, SMS, Razorpay."),
    (15, "SMTP Integration", "WAITING", 0, "SMTP credentials required for email verification, password reset, invoices, and email notifications."),
    (16, "SMS API Integration", "WAITING", 0, "Waiting for SMS provider credentials. WhatsApp is the active OTP channel."),
    (17, "WhatsApp Business API Integration", "NEAR COMPLETE", 95, "Sparklebot integrated: login OTP, payment received, coupons, feedback request, appointment confirm. Templates approved. Final prod smoke pending."),
    (18, "Payment Gateway Integration (Razorpay)", "WAITING", 0, "Waiting for Razorpay merchant account and API keys. App uses UPI QR with manual payment confirmation."),
    (19, "Production Smoke Testing", "IN PROGRESS", 85, "Core modules validated. Final smoke after SMTP, SMS, and Razorpay."),
    (20, "Production Go-Live", "IN PROGRESS", 92, "Application is publicly accessible. Final sign-off depends on SMTP, SMS, and Payment Gateway."),
]

# ===== Sheet 1: Phase / activity status (primary) =====
ws = wb.active
ws.title = "Phase Status"

ws["A1"] = "BillVyapp — Phase / Activity Status"
ws["A1"].font = title_font
ws.merge_cells("A1:E1")

ws["A2"] = "Source: docs/BillVyapp-Project-Status-Report-2026-08-07.md §3 · Report date: 7 August 2026 · Client: The Starr Kuts · Live: https://billvyapp.com"
ws["A2"].font = Font(size=9, color="666666")
ws.merge_cells("A2:E2")

headers = ["#", "Phase / Activity", "Status", "Completion %", "Remarks / Details"]
for c, h in enumerate(headers, 1):
    ws.cell(row=4, column=c, value=h)
style_header(ws, 4, 5)
ws.row_dimensions[4].height = 24

for i, (no, act, status, pct, remarks) in enumerate(phases, 5):
    vals = [no, act, status, f"{pct}%", remarks]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.border = thin
        cell.alignment = wrap if c in (2, 5) else center
    ws.cell(row=i, column=3).fill = fills.get(status, PatternFill())
    ws.cell(row=i, column=3).font = Font(bold=True, size=10)
    ws.row_dimensions[i].height = 48

autosize(ws, [5, 42, 16, 14, 78])
ws.auto_filter.ref = f"A4:E{4 + len(phases)}"
ws.freeze_panes = "A5"
ws.print_title_rows = "4:4"

# ===== Sheet 2: Summary (optional context) =====
ws2 = wb.create_sheet("Summary")
ws2["A1"] = "BillVyapp — Project Status Report"
ws2["A1"].font = title_font
ws2.merge_cells("A1:B1")
meta = [
    ("Client", "The Starr Kuts"),
    ("Application", "BillVyapp"),
    ("Production URL", "https://billvyapp.com"),
    ("Report date", "7 August 2026"),
    ("Git baseline", "rbac (fa4e7bf)"),
    ("Production HEAD", "production/main (cb77277)"),
]
ws2["A3"] = "Field"
ws2["B3"] = "Detail"
style_header(ws2, 3, 2)
for i, (k, v) in enumerate(meta, 4):
    ws2.cell(row=i, column=1, value=k).border = thin
    ws2.cell(row=i, column=2, value=v).border = thin
    ws2.cell(row=i, column=1).font = Font(bold=True)
autosize(ws2, [28, 50])

out = r"D:\BillVyapp\BillVyapp\docs\BillVyapp-Project-Status-Report-2026-08-07.xlsx"
wb.save(out)
print(f"Saved {out} with {len(phases)} phase rows")
